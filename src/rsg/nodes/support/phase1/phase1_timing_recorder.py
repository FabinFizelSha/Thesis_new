"""Async Excel timing/debug recorder for Phase 1 nodes."""

from __future__ import annotations

from pathlib import Path
from threading import Lock, Thread
from typing import Any, Dict, List, Optional, Tuple


class Phase1TimingRecorder:
    """Store Phase 1 timing events and write them to Excel outside the hot path."""

    def __init__(self, enabled: bool, output_path: str, autosave_every: int, logger: Any, sheet_name: str) -> None:
        self.enabled = enabled
        self.output_path = Path(output_path).expanduser().resolve()
        self.autosave_every = max(0, int(autosave_every))
        self.logger = logger
        self.sheet_name = self._sanitize_sheet_name(sheet_name)
        self.samples: List[Dict[str, Any]] = []
        self._lock = Lock()
        self._save_thread: Optional[Thread] = None
        self._save_pending = False
        self._last_saved_count = 0

        if self.enabled:
            self.logger.info(f"Phase 1 timing Excel recorder enabled. Output: {self.output_path}")
        else:
            self.logger.info("Phase 1 timing Excel recorder disabled.")

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        invalid_chars = "[]:*?/\\"
        cleaned = "".join("_" if char in invalid_chars else char for char in name)
        return (cleaned.strip() or "phase1_timing")[:31]

    def add_sample(self, **sample: Any) -> None:
        """Add one event. Accepts arbitrary columns to support both Phase 1 nodes."""
        if not self.enabled:
            return
        with self._lock:
            self.samples.append(dict(sample))
            count = len(self.samples)
        if self.autosave_every > 0 and count % self.autosave_every == 0:
            self.request_save_async()

    def request_save_async(self) -> None:
        """Run the `request save async` operation."""
        if not self.enabled:
            return
        with self._lock:
            if self._save_thread is not None and self._save_thread.is_alive():
                self._save_pending = True
                return
            self._save_pending = False
            self._save_thread = Thread(target=self._save_worker, daemon=True)
            self._save_thread.start()

    def _save_worker(self) -> None:
        while True:
            snapshot = self._snapshot_for_save()
            if snapshot is not None:
                samples, count = snapshot
                self._write_workbook(samples, count)
            with self._lock:
                if self._save_pending:
                    self._save_pending = False
                    continue
                return

    def _snapshot_for_save(self) -> Optional[Tuple[List[Dict[str, Any]], int]]:
        with self._lock:
            count = len(self.samples)
            if count == 0:
                return None
            if count == self._last_saved_count and self.output_path.exists():
                return None
            return list(self.samples), count

    def save(self) -> None:
        """Synchronously write latest timing data. Intended for node shutdown."""
        if not self.enabled:
            return
        thread = self._save_thread
        if thread is not None and thread.is_alive():
            thread.join()
        snapshot = self._snapshot_for_save()
        if snapshot is None:
            return
        samples, count = snapshot
        self._write_workbook(samples, count)

    def _write_workbook(self, samples: List[Dict[str, Any]], count: int) -> None:
        if not self.enabled or not samples:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            self.logger.error("openpyxl is missing; cannot write Phase 1 timing workbook.")
            self.logger.error(str(exc))
            return

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        headers: List[str] = []
        for sample in samples:
            for key in sample.keys():
                if key not in headers:
                    headers.append(key)

        wb = Workbook()
        sheet = wb.active
        sheet.title = self.sheet_name
        sheet.append(headers)
        for sample in samples:
            sheet.append([sample.get(header) for header in headers])

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        fail_fill = PatternFill("solid", fgColor="F4CCCC")
        ok_fill = PatternFill("solid", fgColor="D9EAD3")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        status_col = headers.index("status") + 1 if "status" in headers else None
        if status_col is not None:
            for row in sheet.iter_rows(min_row=2, max_row=len(samples) + 1):
                status = str(row[status_col - 1].value).lower() if row[status_col - 1].value is not None else ""
                if status in {"failed", "dropped", "stale", "error"}:
                    for cell in row:
                        cell.fill = fail_fill
                elif status in {"ok", "published", "sent_to_hydra", "vlm_done"}:
                    row[status_col - 1].fill = ok_fill

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for col in sheet.columns:
            letter = col[0].column_letter
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            sheet.column_dimensions[letter].width = min(max(12, max_len + 2), 50)

        wb.save(self.output_path)
        self._last_saved_count = count
        self.logger.info(f"Saved Phase 1 timing workbook with {count} rows: {self.output_path}")
