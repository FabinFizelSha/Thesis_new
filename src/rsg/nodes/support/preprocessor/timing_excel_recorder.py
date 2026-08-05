"""Timing/debug Excel recorder for the RSG preprocessor.

The important design rule is that Excel generation must not run inside the
real-time frame callback. Building an XLSX workbook with openpyxl can take much
longer than a normal frame, especially after many rows have accumulated. This
recorder therefore stores timing rows cheaply in memory and performs autosaves
on a background thread. A final synchronous save is still performed during node
shutdown.
"""

from __future__ import annotations

from pathlib import Path
from threading import Lock, Thread
from typing import Any, Dict, List, Optional, Tuple


class TimingExcelRecorder:
    """Collect preprocessing timing/debug events and export an Excel report.

    The recorder is inactive when debug mode is disabled. It stores one row for
    each accepted or rejected RGB-D synchronization event. Published frames are
    marked as ``published`` and rejected frames are marked as ``rejected`` with
    the rejection reason highlighted in the generated workbook.

    Autosave is asynchronous so the ROS frame callback is not blocked by Excel
    writing. This avoids occasional long output gaps that can disturb downstream
    processing.
    """

    def __init__(
        self,
        enabled: bool,
        output_path: str,
        autosave_every: int,
        logger: Any,
        sheet_name: str = "timing_debug",
    ) -> None:
        self.enabled = enabled
        self.output_path = Path(output_path).expanduser()
        self.autosave_every = max(0, int(autosave_every))
        self.logger = logger
        self.samples: List[Dict[str, Any]] = []
        self._last_saved_count = 0
        self.sheet_name = self._sanitize_sheet_name(sheet_name)

        self._lock = Lock()
        self._save_thread: Optional[Thread] = None
        self._save_pending = False

        if self.enabled:
            self.logger.info(f"Timing Excel recorder enabled. Output: {self.output_path}")
            self.logger.info("Timing Excel autosave uses a background thread and will not block frame publishing.")
        else:
            self.logger.info("Timing Excel recorder disabled.")

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """Return an Excel-safe worksheet name limited to 31 characters."""
        invalid_chars = "[]:*?/\\"
        cleaned = "".join("_" if char in invalid_chars else char for char in name)
        return (cleaned.strip() or "timing_debug")[:31]

    def add_sample(
        self,
        sequence: int,
        frame_id: str,
        rgb_time: float,
        processing_delay_ms: float,
        rgb_depth_dt_sec: Optional[float],
        rgb_odom_dt_sec: Optional[float],
        odom_status: str,
        status: str = "published",
        reason: str = "ok",
        invalid_depth_ratio: Optional[float] = None,
        imu_status: Optional[str] = None,
        rgb_imu_dt_sec: Optional[float] = None,
    ) -> None:
        """Add one preprocessing timing/debug event.

        This method is intentionally lightweight because it is called from the
        frame-processing callback. It appends one dictionary to memory and, when
        required, starts or requests a background autosave.
        """
        if not self.enabled:
            return

        sample = {
            "sequence": int(sequence),
            "frame_id": frame_id,
            "status": status,
            "reason": reason,
            "rgb_time_sec": float(rgb_time),
            "processing_delay_ms": float(processing_delay_ms),
            "rgb_depth_dt_sec": None if rgb_depth_dt_sec is None else float(rgb_depth_dt_sec),
            "rgb_odom_dt_sec": None if rgb_odom_dt_sec is None else float(rgb_odom_dt_sec),
            "odom_status": odom_status,
            "invalid_depth_ratio": None if invalid_depth_ratio is None else float(invalid_depth_ratio),
            "imu_status": imu_status,
            "rgb_imu_dt_sec": None if rgb_imu_dt_sec is None else float(rgb_imu_dt_sec),
        }

        with self._lock:
            self.samples.append(sample)
            sample_count = len(self.samples)

        # Do not save on the very first frame. Creating an XLSX file for the
        # first sample caused visible frame jitter during startup. Save only at
        # configured intervals and at shutdown.
        if self.autosave_every > 0 and sample_count % self.autosave_every == 0:
            self.request_save_async()

    def request_save_async(self) -> None:
        """Request a non-blocking autosave.

        If a save is already running, mark another save as pending. The worker
        will then write the newest snapshot once the current save finishes.
        """
        if not self.enabled:
            return

        with self._lock:
            if self._save_thread is not None and self._save_thread.is_alive():
                self._save_pending = True
                return
            self._save_pending = False
            self._save_thread = Thread(target=self._save_worker, daemon=True)
            self._save_thread.start()

    def _save_worker(self) -> None:
        """Background autosave loop."""
        while True:
            snapshot = self._snapshot_for_save()
            if snapshot is not None:
                samples_snapshot, saved_count = snapshot
                self._write_workbook(samples_snapshot, saved_count)

            with self._lock:
                if self._save_pending:
                    self._save_pending = False
                    continue
                return

    def _snapshot_for_save(self) -> Optional[Tuple[List[Dict[str, Any]], int]]:
        """Return a copy of all samples if there is new data to save."""
        with self._lock:
            sample_count = len(self.samples)
            if sample_count == 0:
                return None
            if self._last_saved_count == sample_count and self.output_path.exists():
                return None
            return list(self.samples), sample_count

    def save(self) -> None:
        """Synchronously write the latest timing/debug table to Excel.

        This is used at node shutdown. It may block briefly, but it is outside
        normal frame processing.
        """
        if not self.enabled:
            return

        thread = self._save_thread
        if thread is not None and thread.is_alive():
            thread.join()

        snapshot = self._snapshot_for_save()
        if snapshot is None:
            return
        samples_snapshot, saved_count = snapshot
        self._write_workbook(samples_snapshot, saved_count)

    def _write_workbook(self, samples_snapshot: List[Dict[str, Any]], saved_count: int) -> None:
        """Write a snapshot of timing/debug data to an Excel workbook."""
        if not self.enabled or not samples_snapshot:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            self.logger.error(
                "Could not write timing Excel file because openpyxl is missing. "
                "Install it with: sudo apt install python3-openpyxl or python3 -m pip install openpyxl"
            )
            self.logger.error(str(exc))
            return

        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_name

        headers = [
            "Frame sequence",
            "Frame ID",
            "Status",
            "Reason",
            "RGB timestamp [s]",
            "Preprocessing delay [ms]",
            "RGB-depth dt [s]",
            "RGB-odom dt [s]",
            "Odom status",
            "Invalid depth ratio",
            "IMU status",
            "RGB-IMU dt [s]",
        ]
        sheet.append(headers)

        for sample in samples_snapshot:
            sheet.append([
                sample["sequence"],
                sample["frame_id"],
                sample["status"],
                sample["reason"],
                sample["rgb_time_sec"],
                sample["processing_delay_ms"],
                sample["rgb_depth_dt_sec"],
                sample["rgb_odom_dt_sec"],
                sample["odom_status"],
                sample["invalid_depth_ratio"],
                sample["imu_status"],
                sample["rgb_imu_dt_sec"],
            ])

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        rejected_fill = PatternFill("solid", fgColor="F4CCCC")
        published_fill = PatternFill("solid", fgColor="D9EAD3")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in sheet.iter_rows(min_row=2, max_row=len(samples_snapshot) + 1):
            status_value = str(row[2].value).lower() if row[2].value is not None else ""
            if status_value == "rejected":
                for cell in row:
                    cell.fill = rejected_fill
                row[2].font = Font(bold=True, color="9C0006")
                row[3].font = Font(bold=True, color="9C0006")
            elif status_value == "published":
                row[2].fill = published_fill

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = {
            "A": 16,
            "B": 42,
            "C": 14,
            "D": 34,
            "E": 18,
            "F": 24,
            "G": 18,
            "H": 18,
            "I": 26,
            "J": 20,
            "K": 22,
            "L": 18,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=12):
            for cell in row:
                cell.number_format = "0.000000"
        for cell in sheet["F"][1:]:
            cell.number_format = "0.000"
        for cell in sheet["J"][1:]:
            cell.number_format = "0.000"

        if len(samples_snapshot) >= 2:
            chart = LineChart()
            chart.title = "Preprocessing delay per RGB frame"
            chart.style = 13
            chart.y_axis.title = "Delay [ms]"
            chart.x_axis.title = "Incoming RGB frame sequence"
            data = Reference(sheet, min_col=6, min_row=1, max_row=len(samples_snapshot) + 1)
            categories = Reference(sheet, min_col=1, min_row=2, max_row=len(samples_snapshot) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 12
            chart.width = 24
            sheet.add_chart(chart, "N2")

        workbook.save(self.output_path)
        with self._lock:
            self._last_saved_count = max(self._last_saved_count, saved_count)
        self.logger.info(f"Wrote preprocessing timing Excel report: {self.output_path}")
